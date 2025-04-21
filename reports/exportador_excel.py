# reports/exportador_excel.py
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import os

def exportar_reporte(fecha, sucursales, obtener_datos_por_sucursal):
    wb = Workbook()
    for sucursal in sucursales:
        sheet = wb.create_sheet(sucursal)
        datos = obtener_datos_por_sucursal(sucursal)

        row = 1
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        sheet.cell(row=row, column=1, value=f"Sucursal: {sucursal}  | Fecha: {fecha}").font = Font(bold=True)
        row += 2

        def render_section(titulo, columnas, datos):
            nonlocal row
            sheet.cell(row=row, column=1, value=titulo).font = Font(bold=True)
            row += 1
            for col_idx, col in enumerate(columnas, 1):
                sheet.cell(row=row, column=col_idx, value=col).font = Font(bold=True)
            row += 1
            for fila in datos:
                for col_idx, celda in enumerate(fila, 1):
                    sheet.cell(row=row, column=col_idx, value=celda)
                row += 1
            row += 2

        render_section("Facturas", ["Forma", "Total"], datos["facturas"])
        render_section("Notas", ["Forma", "Total"], datos["notas"])
        render_section("Cobranza", ["Forma", "Total"], datos["cobranza"])
        render_section("Devoluciones", ["Forma", "Total"], datos["devoluciones"])
        render_section("Corte de Caja", ["Forma", "Importe"], datos["corte"])
        render_section("Resumen", ["Forma", "Corte", "Revisión", "Diferencia"], datos["resumen"])

        for col in range(1, 10):
            sheet.column_dimensions[chr(64 + col)].width = 20

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    filename = f"ReporteGeneral_{fecha.replace('/', '-')}.xlsx"
    wb.save(filename)
    return os.path.abspath(filename)



